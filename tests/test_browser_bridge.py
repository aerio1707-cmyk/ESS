"""單元測試：core/browser_bridge.py 的 JSON in / JSON out 合約（合成假資料）。

先在一般 CPython 下驗證這層邏輯正確，之後 Pyodide 環境呼叫的是同一份程式碼，
不用另外重新驗證比對邏輯本身。
"""

import json

import openpyxl

from core.browser_bridge import list_sheets_json, mapping_preview_json, run_json


def _make_target(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工作表1"
    ws.append(["客戶名稱", "統一編號", "Acc. SE"])
    ws.append(["測試公司A", "10000001", None])
    ws.append(["示範企業B", "10000002", None])
    wb.save(path)


def _make_roster(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Special Acc list"
    ws.append(["統編", "客戶", "Group Name", "Acc. SE", ""])
    ws.append(["", "", "", "Name", ""])
    ws.append(["10000001", "測試公司A", "測試公司", "王小明", ""])
    wb.save(path)


def test_list_sheets_json(tmp_path):
    target_path = tmp_path / "target.xlsx"
    roster_path = tmp_path / "roster.xlsx"
    _make_target(target_path)
    _make_roster(roster_path)

    result = json.loads(list_sheets_json(json.dumps({"target_path": str(target_path), "roster_path": str(roster_path)})))
    assert result["target_sheets"] == ["工作表1"]
    assert result["roster_sheets"] == ["Special Acc list"]


def test_list_sheets_json_error_on_missing_file(tmp_path):
    result = json.loads(
        list_sheets_json(
            json.dumps({"target_path": str(tmp_path / "missing.xlsx"), "roster_path": str(tmp_path / "missing2.xlsx")})
        )
    )
    assert "error" in result


def test_mapping_preview_json_detects_fields(tmp_path):
    target_path = tmp_path / "target.xlsx"
    roster_path = tmp_path / "roster.xlsx"
    _make_target(target_path)
    _make_roster(roster_path)

    payload = {
        "target_path": str(target_path),
        "target_sheet": "工作表1",
        "roster_path": str(roster_path),
        "roster_sheet": "Special Acc list",
    }
    result = json.loads(mapping_preview_json(json.dumps(payload)))

    assert result["target"]["best_guess"]["tax_id"] == 2
    assert result["target"]["best_guess"]["customer_name"] == 1
    assert result["target"]["best_guess"]["acc_se"] == 3
    assert result["roster"]["best_guess"]["acc_se"] == 4


def test_run_json_end_to_end(tmp_path):
    target_path = tmp_path / "target.xlsx"
    roster_path = tmp_path / "roster.xlsx"
    _make_target(target_path)
    _make_roster(roster_path)

    payload = {
        "target_path": str(target_path),
        "roster_path": str(roster_path),
        "target": {
            "sheet_name": "工作表1",
            "header_row": 1,
            "data_start_row": 2,
            "tax_id_col": 2,
            "customer_name_col": 1,
            "acc_se_col": 3,
        },
        "roster": {
            "sheet_name": "Special Acc list",
            "header_row": 1,
            "data_start_row": 3,
            "tax_id_col": 1,
            "customer_name_col": 2,
            "group_name_col": 3,
            "acc_se_col": 4,
        },
        "formula_col": None,
        "output_dir": str(tmp_path / "out"),
        "source_stem": "target",
    }
    result = json.loads(run_json(json.dumps(payload)))

    assert result["stats"]["total"] == 2
    assert result["preview_rows"][0]["matched_acc_se"] == "王小明"
    assert result["preview_rows"][1]["matched_acc_se"] == "N/A"

    primary_path = result["downloads"]["primary"]
    wb_out = openpyxl.load_workbook(primary_path)
    ws_out = wb_out["工作表1"]
    assert ws_out.max_column == 3  # 沒有新增 Match Status 欄
    assert ws_out.cell(row=2, column=3).value == "王小明"


def test_run_json_target_without_tax_id_column(tmp_path):
    """目標檔沒有統編欄時（只選客戶名稱），應該還是能透過 Pass2 名稱比對成功，不因缺欄位而報錯。"""
    target_path = tmp_path / "target.xlsx"
    roster_path = tmp_path / "roster.xlsx"
    _make_target(target_path)
    _make_roster(roster_path)

    payload = {
        "target_path": str(target_path),
        "roster_path": str(roster_path),
        "target": {
            "sheet_name": "工作表1",
            "header_row": 1,
            "data_start_row": 2,
            "tax_id_col": None,
            "customer_name_col": 1,
            "acc_se_col": 3,
        },
        "roster": {
            "sheet_name": "Special Acc list",
            "header_row": 1,
            "data_start_row": 3,
            "tax_id_col": 1,
            "customer_name_col": 2,
            "group_name_col": 3,
            "acc_se_col": 4,
        },
        "formula_col": None,
        "output_dir": str(tmp_path / "out"),
        "source_stem": "target",
    }
    result = json.loads(run_json(json.dumps(payload)))

    assert result["stats"]["total"] == 2
    assert result["preview_rows"][0]["match_pass"] == "名稱比對成功"
    assert result["preview_rows"][0]["matched_acc_se"] == "王小明"
    # 沒選統編欄，不該冒出「統編欄缺失」這種誤導性異常標記
    assert "統編欄缺失" not in result["preview_rows"][0]["anomalies"]


def test_run_json_target_requires_at_least_one_id_field(tmp_path):
    """目標檔統編／客戶名稱兩者都沒選時，仍應該擋下來（跟名冊一樣要有辦法比對）。"""
    target_path = tmp_path / "target.xlsx"
    roster_path = tmp_path / "roster.xlsx"
    _make_target(target_path)
    _make_roster(roster_path)

    payload = {
        "target_path": str(target_path),
        "roster_path": str(roster_path),
        "target": {
            "sheet_name": "工作表1",
            "header_row": 1,
            "data_start_row": 2,
            "tax_id_col": None,
            "customer_name_col": None,
            "acc_se_col": 3,
        },
        "roster": {
            "sheet_name": "Special Acc list",
            "header_row": 1,
            "data_start_row": 3,
            "tax_id_col": 1,
            "customer_name_col": 2,
            "group_name_col": 3,
            "acc_se_col": 4,
        },
        "formula_col": None,
        "output_dir": str(tmp_path / "out"),
        "source_stem": "target",
    }
    result = json.loads(run_json(json.dumps(payload)))
    assert "error" in result


def test_run_json_missing_mapping_returns_error(tmp_path):
    target_path = tmp_path / "target.xlsx"
    roster_path = tmp_path / "roster.xlsx"
    _make_target(target_path)
    _make_roster(roster_path)

    payload = {
        "target_path": str(target_path),
        "roster_path": str(roster_path),
        "target": {},
        "roster": {},
        "formula_col": None,
        "output_dir": str(tmp_path / "out"),
        "source_stem": "target",
    }
    result = json.loads(run_json(json.dumps(payload)))
    assert "error" in result
