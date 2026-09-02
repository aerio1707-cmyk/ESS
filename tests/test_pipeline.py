"""單元測試：core/pipeline.py 的 run_full_pipeline（CLI／scripts/run_sample_validation.py
共用的完整流程），2026-09-01 起補上已回填主檔＋異常明細表輸出，跟 browser-app 的輸出
內容看齊（合成假資料，不含真實客戶資訊）。
"""

import openpyxl

from core.pipeline import run_full_pipeline


def _make_target(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工作表1"
    ws.append(["客戶名稱", "統一編號"])  # 刻意不放 Acc. SE 欄，一併驗證自動新增空白欄
    ws.append(["測試公司A", "10000001"])
    wb.save(path)


def _make_roster(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Special Acc list"
    ws.append(["統編", "客戶", "Group Name", "Acc. SE", ""])
    ws.append(["", "", "", "Name", ""])
    ws.append(["10000001", "測試公司A", "測試公司", "王小明", ""])
    wb.save(path)


def test_run_full_pipeline_writes_primary_and_unmatched_outputs_like_browser_app(tmp_path):
    target_path = tmp_path / "target.xlsx"
    roster_path = tmp_path / "roster.xlsx"
    _make_target(target_path)
    _make_roster(roster_path)

    result = run_full_pipeline(
        target_path=target_path,
        target_sheet="工作表1",
        roster_path=roster_path,
        roster_sheet="Special Acc list",
        output_dir=tmp_path / "out",
        interactive=False,
    )

    assert result.stats.total == 1
    assert result.primary_output_path.exists()
    assert result.unmatched_report_path.exists()
    assert result.review_csv_path.exists()

    wb_out = openpyxl.load_workbook(result.primary_output_path)
    ws_out = wb_out["工作表1"]
    # 目標檔沒有 Acc. SE 欄，應該自動用資料結束後的空白欄（第3欄），並標上表頭＋黃色底色
    assert ws_out.cell(row=1, column=3).value == "Acc. SE"
    assert ws_out.cell(row=1, column=3).fill.fgColor.rgb == "00FFFF00"
    assert ws_out.cell(row=2, column=3).value == "王小明"
