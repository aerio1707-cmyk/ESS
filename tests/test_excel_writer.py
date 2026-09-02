"""單元測試：core/excel_writer.py（合成假資料，不含真實客戶資訊）。"""

import openpyxl
from openpyxl.styles import PatternFill

from core.excel_writer import write_primary_output, write_unmatched_report
from core.models import AnomalyCode, FieldMapping, MatchPass, MatchResult

PRESET_FILL = PatternFill("solid", fgColor="FFF2CC")  # 原始檔案本來就有的自訂底色


def _make_target_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["客戶名稱", "統一編號", "Acc. SE"])
    ws.append(["測試公司A", "10000001", "=VLOOKUP(B2,X:X,1,0)"])  # 模擬既有公式欄
    ws.append(["測試公司B", "10000002", None])
    ws.append(["測試公司C", "10000003", None])
    # 模擬原始檔案本身就有的格式/底色（例如使用者手動標記過的儲存格），驗證不會被動到
    ws.cell(row=1, column=1).fill = PRESET_FILL
    ws.cell(row=3, column=3).fill = PRESET_FILL
    wb.save(path)


def _results():
    return [
        MatchResult(
            row_index=2,
            raw_tax_id="10000001",
            raw_customer_name="測試公司A",
            clean_tax_id="10000001",
            clean_customer_name="測試公司A",
            matched_acc_se="王小明",
            match_pass=MatchPass.TAX_ID,
        ),
        MatchResult(
            row_index=3,
            raw_tax_id="10000002",
            raw_customer_name="測試公司B",
            clean_tax_id="10000002",
            clean_customer_name="測試公司B",
            matched_acc_se="甲負責人",
            match_pass=MatchPass.TAX_ID,
            anomalies=[AnomalyCode.ROSTER_DUPLICATE_CONFLICT],
            duplicate_candidates=["甲負責人", "乙負責人"],
        ),
        MatchResult(
            row_index=4,
            raw_tax_id="10000003",
            raw_customer_name="測試公司C",
            clean_tax_id="10000003",
            clean_customer_name="測試公司C",
            matched_acc_se="N/A",
            match_pass=MatchPass.UNMATCHED,
        ),
    ]


def test_write_primary_output_only_overwrites_acc_se_values(tmp_path):
    target_path = tmp_path / "target.xlsx"
    _make_target_workbook(target_path)

    mapping = FieldMapping(
        sheet_name="Sheet",
        header_row=1,
        data_start_row=2,
        tax_id_col=2,
        customer_name_col=1,
        acc_se_output_col=3,
    )

    out_path = write_primary_output(target_path, mapping, _results(), output_dir=tmp_path / "out")

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    # 不新增任何欄位
    assert ws.max_column == 3
    # 原本是公式，回填後應為純值，不再是公式字串
    assert ws.cell(row=2, column=3).value == "王小明"
    assert ws.cell(row=3, column=3).value == "甲負責人"
    assert ws.cell(row=4, column=3).value == "N/A"
    # 統編/客戶名稱欄完全不變
    assert ws.cell(row=2, column=1).value == "測試公司A"
    assert ws.cell(row=2, column=2).value == "10000001"
    # 不新增凍結窗格
    assert ws.freeze_panes is None


def test_write_primary_output_preserves_original_formatting(tmp_path):
    target_path = tmp_path / "target.xlsx"
    _make_target_workbook(target_path)
    mapping = FieldMapping(
        sheet_name="Sheet", header_row=1, data_start_row=2, tax_id_col=2, customer_name_col=1, acc_se_output_col=3
    )

    out_path = write_primary_output(target_path, mapping, _results(), output_dir=tmp_path / "out")
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    # 原始檔案本來就有的底色（含 Acc.SE 儲存格自己的底色）維持原樣，不被覆寫/清除
    assert ws.cell(row=1, column=1).fill.fgColor.rgb == "00FFF2CC"
    assert ws.cell(row=3, column=3).fill.fgColor.rgb == "00FFF2CC"
    # 無法匹配/重複衝突等狀態完全不上色（不套用任何新樣式）
    assert ws.cell(row=4, column=3).fill.fgColor.rgb in (None, "00000000")
    assert ws.cell(row=2, column=3).fill.fgColor.rgb in (None, "00000000")


def test_write_primary_output_does_not_touch_existing_acc_se_header(tmp_path):
    """目標檔本來就有「Acc. SE」表頭文字時，不該被改寫或上色（維持既有規則）。"""
    target_path = tmp_path / "target.xlsx"
    _make_target_workbook(target_path)
    mapping = FieldMapping(
        sheet_name="Sheet", header_row=1, data_start_row=2, tax_id_col=2, customer_name_col=1, acc_se_output_col=3
    )

    out_path = write_primary_output(target_path, mapping, _results(), output_dir=tmp_path / "out")
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    assert ws.cell(row=1, column=3).value == "Acc. SE"  # 文字沒被改動
    assert ws.cell(row=1, column=3).fill.fgColor.rgb in (None, "00000000")  # 沒被上色


def test_write_primary_output_labels_and_colors_new_blank_acc_se_column(tmp_path):
    """目標檔沒有 Acc. SE 欄、改用資料結束後緊接的空白欄時，該欄表頭要自動寫入文字＋黃色底色。"""
    target_path = tmp_path / "target.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["客戶名稱", "統一編號"])  # 完全沒有 Acc. SE 欄
    ws.append(["測試公司A", "10000001"])
    wb.save(target_path)

    mapping = FieldMapping(
        sheet_name="Sheet", header_row=1, data_start_row=2, tax_id_col=2, customer_name_col=1, acc_se_output_col=3
    )
    results = [
        MatchResult(
            row_index=2,
            raw_tax_id="10000001",
            raw_customer_name="測試公司A",
            clean_tax_id="10000001",
            clean_customer_name="測試公司A",
            matched_acc_se="王小明",
            match_pass=MatchPass.TAX_ID,
        )
    ]

    out_path = write_primary_output(target_path, mapping, results, output_dir=tmp_path / "out")
    wb_out = openpyxl.load_workbook(out_path)
    ws_out = wb_out.active

    assert ws_out.cell(row=1, column=3).value == "Acc. SE"
    assert ws_out.cell(row=1, column=3).fill.fgColor.rgb == "00FFFF00"
    assert ws_out.cell(row=2, column=3).value == "王小明"


def test_write_unmatched_report_only_includes_unmatched_and_duplicate(tmp_path):
    out_path = write_unmatched_report(_results(), "target", output_dir=tmp_path / "out")
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    rows = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert "測試公司B" in rows  # 重複衝突
    assert "測試公司C" in rows  # 無法匹配
    assert "測試公司A" not in rows  # 正常命中，不應出現


def test_write_unmatched_report_includes_near_miss_and_variant_normalized_rows(tmp_path):
    results = [
        MatchResult(
            row_index=5,
            raw_tax_id="",
            raw_customer_name="無法匹配公司",
            clean_tax_id="",
            clean_customer_name="無法匹配公司",
            matched_acc_se="N/A",
            match_pass=MatchPass.UNMATCHED,
            near_miss_customer_name="臺灣恩益禧股份有限公司",
            near_miss_group_name=None,
        ),
        MatchResult(
            row_index=6,
            raw_tax_id="",
            raw_customer_name="台灣恩益禧股份有限公司",
            clean_tax_id="",
            clean_customer_name="台灣恩益禧股份有限公司",
            matched_acc_se="施柏屹",
            match_pass=MatchPass.CUSTOMER_NAME,
            anomalies=[AnomalyCode.VARIANT_CHAR_NORMALIZED],
        ),
    ]
    out_path = write_unmatched_report(results, "target", output_dir=tmp_path / "out")
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    rows_by_customer = {ws.cell(row=r, column=2).value: r for r in range(2, ws.max_row + 1)}

    unmatched_row = rows_by_customer["無法匹配公司"]
    assert ws.cell(row=unmatched_row, column=7).value == "臺灣恩益禧股份有限公司"  # 疑似異體字候選(客戶名稱)
    assert not ws.cell(row=unmatched_row, column=8).value  # 疑似異體字候選(Group Name)：沒有候選則空白

    # 已透過異體字正規化才比對成功的列，也該被收進報表方便稽核，即使它其實比對成功了
    variant_row = rows_by_customer["台灣恩益禧股份有限公司"]
    assert ws.cell(row=variant_row, column=5).value == AnomalyCode.VARIANT_CHAR_NORMALIZED.value
    assert not ws.cell(row=variant_row, column=7).value  # 已成功比對，不需要近似候選
