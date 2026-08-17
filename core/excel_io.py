"""Excel 讀取、標頭偵測、雙層標頭解析、合併儲存格回填。

規劃書 2.4：正式比對邏輯不可依賴 Excel 公式快取值，一律用 data_only=True 只把
快取值當「參考顯示」用（例如顯示原公式判定 N/A 筆數），實際比對用程式重新計算。
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def list_sheets(path: str | Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def load_worksheet(path: str | Path, sheet_name: str) -> Worksheet:
    """讀取指定工作表（非 read_only，才能取得 merged_cells 範圍）。"""
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    return wb[sheet_name]


def build_merged_value_map(ws: Worksheet) -> dict[tuple[int, int], object]:
    """回傳 {(row, col): 值} 供合併儲存格範圍內的非首格查值使用（規劃書 3.2 #12）。"""
    merged_map: dict[tuple[int, int], object] = {}
    for merged_range in ws.merged_cells.ranges:
        top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if row == merged_range.min_row and col == merged_range.min_col:
                    continue
                merged_map[(row, col)] = top_left
    return merged_map


def cell_value(ws: Worksheet, row: int, col: int, merged_map: dict[tuple[int, int], object]) -> object:
    value = ws.cell(row=row, column=col).value
    if value is None and (row, col) in merged_map:
        return merged_map[(row, col)]
    return value


def row_values(
    ws: Worksheet, row: int, max_col: int, merged_map: dict[tuple[int, int], object]
) -> list[object]:
    return [cell_value(ws, row, col, merged_map) for col in range(1, max_col + 1)]


def detect_header_row(ws: Worksheet, merged_map: dict[tuple[int, int], object], max_scan: int = 5) -> int:
    """掃描前 max_scan 列，取「非空白比例最高且內容多為文字」的那一列（規劃書 2.2）。"""
    max_col = ws.max_column
    best_row, best_score = 1, -1.0
    for row in range(1, min(max_scan, ws.max_row) + 1):
        values = row_values(ws, row, max_col, merged_map)
        non_empty = [v for v in values if v is not None and str(v).strip() != ""]
        if not values:
            continue
        non_empty_ratio = len(non_empty) / len(values)
        if not non_empty:
            continue
        text_count = sum(1 for v in non_empty if isinstance(v, str))
        text_ratio = text_count / len(non_empty)
        score = non_empty_ratio + text_ratio
        if score > best_score:
            best_row, best_score = row, score
    return best_row


def parse_single_header(
    ws: Worksheet, header_row: int, merged_map: dict[tuple[int, int], object]
) -> dict[str, int]:
    """單層標頭：{標頭文字: 欄號}。同名欄後出現者覆蓋前者對應到最後一欄（罕見情況才會發生）。"""
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = cell_value(ws, header_row, col, merged_map)
        if value is None:
            continue
        text = str(value).strip()
        if text:
            mapping[text] = col
    return mapping


def parse_double_header(
    ws: Worksheet,
    group_row: int,
    sub_row: int,
    merged_map: dict[tuple[int, int], object],
) -> dict[str, int]:
    """雙層標頭（規劃書 2.5）：group_row 為群組名（可重複），sub_row 為子欄名。

    回傳鍵格式：有子欄名的欄位用 "群組__子欄"；子欄名為空的欄位（如統編/客戶/Group Name
    本身沒有子欄）直接用群組名。
    """
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        group_value = cell_value(ws, group_row, col, merged_map)
        sub_value = cell_value(ws, sub_row, col, merged_map)
        group_text = str(group_value).strip() if group_value is not None else ""
        sub_text = str(sub_value).strip() if sub_value is not None else ""
        if not group_text and not sub_text:
            continue
        key = f"{group_text}__{sub_text}" if sub_text else group_text
        if key not in mapping:
            mapping[key] = col
    return mapping
