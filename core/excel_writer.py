"""模組 6/7：Excel 自動格式化與排版輸出、異常明細表（規劃書第七、八章）。"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.models import AnomalyCode, FieldMapping, MatchPass, MatchResult

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

THIN_GRAY = Side(style="thin", color="D9D9D9")
CELL_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)

MIN_COL_WIDTH = 10
MAX_COL_WIDTH = 50
BASE_ROW_HEIGHT = 20

NEW_ACC_SE_HEADER_TEXT = "Acc. SE"
NEW_ACC_SE_HEADER_FILL = PatternFill("solid", fgColor="FFFF00")

ANOMALY_ADVICE: dict[str, str] = {
    MatchPass.UNMATCHED.value: "名冊查無此客戶，如為正式客戶請確認是否應補進名冊",
    AnomalyCode.NON_STANDARD_TAX_ID_CODE.value: "暫編代碼，待補正式統編後再重新比對",
    AnomalyCode.PERSONAL_ID.value: "個人戶（身分證字號格式），可能本來就無需指定 Acc. SE",
    AnomalyCode.MISSING_TAX_ID.value: "來源統編欄缺值，請確認原始資料",
    AnomalyCode.MISSING_CUSTOMER_NAME.value: "來源客戶名稱欄缺值，請確認原始資料",
    AnomalyCode.FORMULA_ERROR.value: "來源欄位為公式錯誤值，請確認原始資料",
    AnomalyCode.ROSTER_DUPLICATE_CONFLICT.value: "名冊同統編/名稱對應多位負責人，請人工確認正確人選",
    AnomalyCode.VARIANT_CHAR_NORMALIZED.value: "已透過異體字對照表正規化才比對成功，建議快速確認回填結果是否正確",
}


def _apply_borders(ws: Worksheet, header_row: int, max_row: int, max_col: int) -> None:
    for row in range(header_row, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = CELL_BORDER


def _autosize_columns(ws: Worksheet, header_row: int, max_row: int, max_col: int) -> None:
    """自動欄寬（10~50），超過上限者對「資料列」開啟 wrap_text（標頭列維持置中不換行）。"""
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        longest = 0
        for row in range(header_row, max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                longest = max(longest, len(str(value)))
        width = min(max(MIN_COL_WIDTH, longest + 2), MAX_COL_WIDTH)
        ws.column_dimensions[letter].width = width

        if longest + 2 > MAX_COL_WIDTH:
            for row in range(header_row + 1, max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")


def _autosize_rows(ws: Worksheet, header_row: int, max_row: int, max_col: int) -> None:
    """依換行數/預估 wrap 後行數動態設定列高（基準 20pt，多行按比例增加）。"""
    for row in range(header_row, max_row + 1):
        max_lines = 1
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            text = str(cell.value)
            lines = text.count("\n") + 1
            if cell.alignment and cell.alignment.wrap_text:
                col_width = ws.column_dimensions[get_column_letter(col)].width or MAX_COL_WIDTH
                lines = max(lines, -(-len(text) // max(int(col_width), 1)))
            max_lines = max(max_lines, lines)
        ws.row_dimensions[row].height = BASE_ROW_HEIGHT * max_lines


def write_primary_output(
    target_path: str | Path,
    mapping: FieldMapping,
    results: list[MatchResult],
    output_dir: str | Path = "output",
    source_stem: str | None = None,
) -> Path:
    """回填主檔：只覆寫 Acc. SE 欄的值，其餘欄位、格式、儲存格顏色完全維持原始檔案不變。

    不新增任何欄位（如 Match Status）、不上色、不調整欄寬列高/凍結窗格/框線——
    使用者 2026-08-18 明確要求輸出檔案除了回填欄位本身，其他一律不動（含既有格式）。

    唯一例外（2026-09-01）：如果目標檔本來就沒有 Acc. SE 欄，欄位對應時會自動選用資料
    結束後緊接的空白欄（見 core/field_mapping.py 的 acc_se_new_column）。這種情況下，那一欄
    的表頭列本來就是空的，此時才會額外寫入「Acc. SE」文字並上黃色底色，方便收到檔案的人
    看得懂這一欄是什麼；如果選到的欄位表頭本來就有文字（既有欄位），則完全不動，維持原規則。
    """
    wb = load_workbook(target_path)
    ws = wb[mapping.sheet_name]

    header_cell = ws.cell(row=mapping.header_row, column=mapping.acc_se_output_col)
    if header_cell.value in (None, ""):
        header_cell.value = NEW_ACC_SE_HEADER_TEXT
        header_cell.fill = NEW_ACC_SE_HEADER_FILL

    results_by_row = {r.row_index: r for r in results}

    for row in range(mapping.data_start_row, ws.max_row + 1):
        result = results_by_row.get(row)
        if result is None:
            continue
        ws.cell(row=row, column=mapping.acc_se_output_col, value=result.matched_acc_se)

    stem = source_stem or Path(target_path).stem
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"{stem}_已回填_{date.today():%Y%m%d}.xlsx"
    wb.save(out_path)
    return out_path


UNMATCHED_REPORT_HEADERS = [
    "原始列號",
    "客戶名稱",
    "統編(原始值)",
    "統編(清洗後)",
    "失敗原因分類",
    "建議處理方式",
    "疑似異體字候選(客戶名稱)",
    "疑似異體字候選(Group Name)",
]

_REVIEW_ANOMALIES = {AnomalyCode.ROSTER_DUPLICATE_CONFLICT, AnomalyCode.VARIANT_CHAR_NORMALIZED}


def _needs_report(result: MatchResult) -> bool:
    if result.match_pass == MatchPass.UNMATCHED:
        return True
    return bool(_REVIEW_ANOMALIES & set(result.anomalies))


def _failure_reason(result: MatchResult) -> str:
    if AnomalyCode.ROSTER_DUPLICATE_CONFLICT in result.anomalies:
        return AnomalyCode.ROSTER_DUPLICATE_CONFLICT.value
    if result.anomalies:
        return result.anomalies[0].value
    return result.match_pass.value


def write_unmatched_report(
    results: list[MatchResult],
    source_stem: str,
    output_dir: str | Path = "output",
) -> Path:
    """未匹配 / 異常資料明細表（規劃書 6.2、8.1）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "異常明細"

    for col, header in enumerate(UNMATCHED_REPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
    ws.freeze_panes = "A2"

    row_idx = 2
    for result in results:
        if not _needs_report(result):
            continue
        reason = _failure_reason(result)
        ws.cell(row=row_idx, column=1, value=result.row_index)
        ws.cell(row=row_idx, column=2, value=result.raw_customer_name)
        ws.cell(row=row_idx, column=3, value=result.raw_tax_id)
        ws.cell(row=row_idx, column=4, value=result.clean_tax_id)
        ws.cell(row=row_idx, column=5, value=reason)
        ws.cell(row=row_idx, column=6, value=ANOMALY_ADVICE.get(reason, "請人工複核"))
        ws.cell(row=row_idx, column=7, value=result.near_miss_customer_name or "")
        ws.cell(row=row_idx, column=8, value=result.near_miss_group_name or "")
        row_idx += 1

    max_row = row_idx - 1
    max_col = len(UNMATCHED_REPORT_HEADERS)
    if max_row >= 1:
        _apply_borders(ws, 1, max_row, max_col)
        _autosize_columns(ws, 1, max_row, max_col)
        _autosize_rows(ws, 1, max_row, max_col)

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"{source_stem}_異常明細_{date.today():%Y%m%d}.xlsx"
    wb.save(out_path)
    return out_path
